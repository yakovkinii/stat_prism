#  Copyright (C) 2023-2026  StatPrism Team
#  Balashevych A. K., Petrova N. V., Yakovkin I. I.
#
#  This file is part of StatPrism.
#
#  StatPrism is free software: you can redistribute it and/or modify it under
#  the terms of the GNU General Public License as published by the Free Software
#  Foundation, either version 3 of the License, or (at your option) any later
#  version.
#
#  StatPrism is distributed in the hope that it will be useful, but WITHOUT ANY
#  WARRANTY; without even the implied warranty of MERCHANTABILITY or FITNESS FOR
#  A PARTICULAR PURPOSE.  See the GNU General Public License for more details.
#
#  You should have received a copy of the GNU General Public License along with
#  StatPrism.  If not, see <https://www.gnu.org/licenses/>.

# VALIDATED

import logging

import pandas as pd
from openpyxl.styles import PatternFill
from PySide6.QtWidgets import QFileDialog

from src.common.constant import hex_to_argb


def export_data_to_excel(parent_widget, data):
    """Prompt for a path and write `data` to .xlsx, painting each header cell with its
    column's color tag. Shared by the Raw Data and data-processing result cards."""
    if data is None or data.n_columns() == 0:
        logging.info("No data to export")
        return

    file_path, _ = QFileDialog.getSaveFileName(parent_widget, "Export to Excel", "", "Excel files (*.xlsx)")
    if not file_path:
        return
    if not file_path.endswith(".xlsx"):
        file_path += ".xlsx"

    try:
        df = data.get_dataframe()
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Sheet1")
            worksheet = writer.sheets["Sheet1"]
            # Paint each header cell with its column's color tag (row 1; openpyxl is 1-based).
            for col_index, name in enumerate(df.columns, start=1):
                argb = hex_to_argb(data[name].color)
                if argb:
                    worksheet.cell(row=1, column=col_index).fill = PatternFill(fill_type="solid", fgColor=argb)
    except Exception as e:
        logging.error(f"Failed to export data to Excel: {e}")
